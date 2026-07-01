"""Dump a specific A1-style range of one sheet: COORD | value | formula.

Usage:
    python tools/dump_range.py "<xlsx>" "Sheet Name" A1:F30 [A1:K35 ...]
"""
import sys
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries


def main():
    path, sheet = sys.argv[1], sys.argv[2]
    ranges = sys.argv[3:]
    wb_v = load_workbook(path, data_only=True)
    wb_f = load_workbook(path, data_only=False)
    ws_v, ws_f = wb_v[sheet], wb_f[sheet]
    for rng in ranges:
        print(f"\n----- {sheet}!{rng} -----")
        min_col, min_row, max_col, max_row = range_boundaries(rng)
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cv = ws_v.cell(row=row, column=col)
                cf = ws_f.cell(row=row, column=col)
                val = cv.value
                formula = cf.value if isinstance(cf.value, str) and cf.value.startswith("=") else ""
                if val is None and not formula:
                    continue
                vs = repr(val)
                if len(vs) > 55:
                    vs = vs[:55] + "..."
                line = f"{cv.coordinate:>6} | {vs}"
                if formula:
                    line += f"  | {formula}"
                print(line)


if __name__ == "__main__":
    main()
