"""Check every _Report.xlsx for the N83 = $AM$36 (off-by-one) 85th-percentile bug.

Reports how many report workbooks contain the typo on each Set-up sheet, and
lists any files that do NOT (i.e., a different/older template).
"""
import sys
import warnings

warnings.simplefilter("ignore")
sys.path.insert(0, r"C:\Users\moshanreh\Desktop\Mohammad\Claude\Intersection")

from openpyxl import load_workbook
from traffic_diag.discovery import find_studies

BASE = r"C:\Users\moshanreh\Desktop\Mohammad\Speed and Volume Studies"
SHEETS = ["Merged Set up", "Incoming Set up", "Outgoing Set up"]

studies = [s for s in find_studies(BASE) if s.report_xlsx]
n_total = 0
n_buggy = 0
clean = []
no_cell = []
errors = []

for s in studies:
    n_total += 1
    try:
        wb = load_workbook(s.report_xlsx, data_only=False, read_only=True)
        buggy_sheets = []
        present = False
        for sh in SHEETS:
            if sh in wb.sheetnames:
                present = True
                f = wb[sh]["N83"].value
                if isinstance(f, str) and "$AM$36" in f:
                    buggy_sheets.append(sh)
        wb.close()
        if buggy_sheets:
            n_buggy += 1
        elif present:
            clean.append(s.study_id)
        else:
            no_cell.append(s.study_id)
    except Exception as e:
        errors.append((s.study_id, str(e)[:60]))

print(f"Report workbooks checked : {n_total}")
print(f"Contain the N83=$AM$36 bug: {n_buggy}")
print(f"Set-up sheets but NO bug  : {len(clean)}")
print(f"No Set-up sheet found     : {len(no_cell)}")
print(f"Errors opening            : {len(errors)}")
if clean:
    print("\nClean (different template?):")
    for c in clean[:20]:
        print("  ", c)
if errors:
    print("\nErrors:")
    for sid, e in errors[:10]:
        print("  ", sid, "->", e)
