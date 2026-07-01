"""One-off inspector: dump an existing radar _Report.xlsx so we can replicate it.

Usage:
    python tools/inspect_report.py "<path to _Report.xlsx>" [sheet substring]

Prints, for each matching sheet, every non-empty cell as: COORD | value | formula
so we can reverse-engineer the statistics, tables, and the 7-day selection logic.
"""
import sys
from openpyxl import load_workbook


def dump(path, sheet_filter=None, max_cells=4000):
    wb_v = load_workbook(path, data_only=True, read_only=True)
    wb_f = load_workbook(path, data_only=False, read_only=True)
    print(f"# FILE: {path}")
    print(f"# SHEETS: {wb_v.sheetnames}\n")
    for name in wb_v.sheetnames:
        if sheet_filter and sheet_filter.lower() not in name.lower():
            continue
        ws_v = wb_v[name]
        ws_f = wb_f[name]
        try:
            dims = ws_v.calculate_dimension()
        except Exception:
            dims = "?"
        print(f"\n===== SHEET: {name}  (dims={dims}) =====")
        shown = 0
        for row_v, row_f in zip(ws_v.iter_rows(), ws_f.iter_rows()):
            for cv, cf in zip(row_v, row_f):
                val = cv.value
                formula = cf.value if isinstance(cf.value, str) and cf.value.startswith("=") else ""
                if val is None and not formula:
                    continue
                vs = repr(val)
                if len(vs) > 60:
                    vs = vs[:60] + "..."
                line = f"{cv.coordinate:>5} | {vs}"
                if formula:
                    line += f"  | {formula}"
                print(line)
                shown += 1
                if shown >= max_cells:
                    print(f"... (truncated at {max_cells} cells)")
                    return


if __name__ == "__main__":
    path = sys.argv[1]
    sheet_filter = sys.argv[2] if len(sys.argv) > 2 else None
    dump(path, sheet_filter)
