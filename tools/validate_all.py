"""Efficient single-pass validation of Python output vs the legacy Excel.

For each study: load once, compute the three direction metrics once, open the
workbook once, then check BOTH the scalar stats (Set-up B6..B13) and the full
hourly table (Report rows 31..54: weekday-85th, per-day counts, averages).

Usage:  python tools/validate_all.py [--sample N] [--year YYYY]
Prints flushed progress and a final summary; writes details to validate_results.csv.
"""
import os
import sys
import warnings

warnings.simplefilter("ignore")
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from openpyxl import load_workbook

from traffic_diag.config import (DIRECTION, TS, DEFAULT_ANALYSIS,
                                 DEFAULT_BASE as BASE, REPO_ROOT)
from traffic_diag.discovery import find_studies
from traffic_diag.metrics import DOW_NAMES, compute_direction_metrics
from traffic_diag.study import load_study
from traffic_diag.validate import day_columns, find_hourly_header

SCALAR_CELLS = {"B6": "p85_speed", "B7": "average_speed", "B8": "median_speed",
                "B9": "max_speed", "B10": "adt", "B11": "average_weekday_traffic",
                "B13": "total_vehicles"}
SETUP = {"Merged": "Merged Set up", "Incoming": "Incoming Set up", "Outgoing": "Outgoing Set up"}
REPORT = {"Merged": "Merged Report", "Incoming": "Incoming Report", "Outgoing": "Outgoing Report"}
DAY_COL = {n: 3 + i for i, n in enumerate(DOW_NAMES)}
TOL = 0.01


def summ(m):
    s = m.summary()
    return {"p85_speed": s["p85_speed"], "average_speed": s["average_speed"],
            "median_speed": s["median_speed"], "max_speed": s["max_speed"],
            "adt": s["adt"], "average_weekday_traffic": s["average_weekday_traffic"],
            "total_vehicles": s["total_vehicles"]}


def close(pv, ev):
    if isinstance(ev, str):
        ev = None
    if pv is None and ev is None:
        return True
    if pv is None or ev is None:
        return False
    return abs(float(pv) - float(ev)) <= TOL


def main():
    args = sys.argv[1:]
    year = int(args[args.index("--year") + 1]) if "--year" in args else None
    sample = int(args[args.index("--sample") + 1]) if "--sample" in args else None
    studies = find_studies(BASE, year=year)
    if sample:
        studies = studies[::max(1, len(studies) // sample)][:sample]
    cfg = DEFAULT_ANALYSIS

    s_tot = s_ok = h_tot = h_ok = 0
    rows = []
    errors = 0
    for i, study in enumerate(studies, 1):
        xlsx = study.report_xlsx
        if not xlsx:
            continue
        try:
            sd = load_study(study, cfg=cfg)
            full = sd.raw
            if full.empty:
                continue
            dates = sorted(full[TS].dt.normalize().unique())
            start = dates[0]
            end = dates[-1] + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
            window = full[(full[TS] >= start) & (full[TS] <= end)]
            design = dates[: max(1, cfg.study_days - 2)]
            metrics = {}
            for d in ("Merged", "Incoming", "Outgoing"):
                sub = window if d == "Merged" else window[window[DIRECTION] == d]
                metrics[d] = compute_direction_metrics(sub, d, sd.speed_limit, cfg,
                                                       n_days=len(dates), design_dates=design)
            wb = load_workbook(xlsx, data_only=True)
            for d in ("Merged", "Incoming", "Outgoing"):
                if SETUP[d] in wb.sheetnames:
                    ws = wb[SETUP[d]]
                    got = summ(metrics[d])
                    for cell, metric in SCALAR_CELLS.items():
                        ev = ws[cell].value
                        if isinstance(ev, (int, float)):
                            s_tot += 1
                            ok = close(got[metric], ev)
                            s_ok += ok
                            if not ok:
                                rows.append((study.study_id, d, "scalar", metric, got[metric], ev))
                if REPORT[d] in wb.sheetnames:
                    ws = wb[REPORT[d]]
                    m = metrics[d]
                    hv, p85 = m.hourly_volume, m.hourly_weekday_p85
                    # Locate the hourly table: its header row varies by template
                    # vintage (28, 29 or 30 across the archive).
                    hdr = find_hourly_header(ws)
                    if hdr is None:
                        continue
                    for hh in range(24):
                        r = hdr + 1 + hh
                        pv = None if p85 is None or pd.isna(p85.get(hh)) else float(p85.get(hh))
                        h_tot += 1; ok = close(pv, ws.cell(row=r, column=2).value); h_ok += ok
                        if not ok:
                            rows.append((study.study_id, d, "hourly", f"p85_h{hh}", pv,
                                         ws.cell(row=r, column=2).value))
                        # Excel hourly columns are Day1..Day7 (chronological) with
                        # weekday-NAME headers on the header row — map by name.
                        for col in range(3, 10):
                            day = ws.cell(row=hdr, column=col).value   # e.g. "Saturday"
                            if day not in DAY_COL:
                                continue
                            dcols = day_columns(hv, day)
                            # absent weekday = the study never ran that day = 0
                            pv = float(hv.loc[hh, dcols].sum()) if dcols else 0.0
                            h_tot += 1; ok = close(pv, ws.cell(row=r, column=col).value); h_ok += ok
                            if not ok:
                                rows.append((study.study_id, d, "hourly", f"{day[:3]}_h{hh}", pv,
                                             ws.cell(row=r, column=col).value))
            wb.close()
        except Exception as e:
            errors += 1
            rows.append((study.study_id, "-", "error", "-", str(e)[:80], None))
        if i % 20 == 0:
            print(f"  ...{i}/{len(studies)} studies "
                  f"(scalar {s_ok}/{s_tot}, hourly {h_ok}/{h_tot})", flush=True)

    print(f"\nStudies: {len(studies)}  errors: {errors}", flush=True)
    print(f"SCALAR cells: {s_ok}/{s_tot} = {s_ok/max(1,s_tot)*100:.3f}%", flush=True)
    print(f"HOURLY cells: {h_ok}/{h_tot} = {h_ok/max(1,h_tot)*100:.3f}%", flush=True)
    df = pd.DataFrame(rows, columns=["study", "direction", "kind", "field", "python", "excel"])
    df.to_csv(os.path.join(REPO_ROOT, "validate_results.csv"), index=False)
    mism = df[df.kind != "error"]
    print(f"Mismatched cells: {len(mism)} across {mism['study'].nunique()} studies", flush=True)
    if len(mism):
        print(mism.head(25).to_string(index=False), flush=True)


if __name__ == "__main__":
    main()
