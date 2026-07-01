"""Locate every mis-referenced cell in the per-hour percentile sub-tables.

Each per-hour cumulative column (the running-sum cells of the form
``=IFERROR(IF($COL$ROW=0, <prev>, $COL$ROW+<prev>), "*")``) must reference a
SINGLE row — that hour's row in the per-hour speed-count block. Any cell whose
absolute $COL$ROW reference disagrees with the column's modal row is a template
bug (the cause of the wrong midnight / late-night Weekday 85th percentile).
"""
import re
import sys
import warnings
from collections import Counter, defaultdict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

warnings.simplefilter("ignore")

ABS = re.compile(r"\$([A-Z]+)\$(\d+)")
CUM = re.compile(r"IF\(\$[A-Z]+\$\d+=0")   # the running-sum cumulative pattern


def scan(path, sheets):
    wb = load_workbook(path, data_only=False)
    for sheet in sheets:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        by_col = defaultdict(list)   # column letter -> [(coord, ref_row)]
        for row in ws.iter_rows():
            for c in row:
                f = c.value
                if isinstance(f, str) and CUM.search(f):
                    m = ABS.search(f)
                    if m:
                        by_col[c.column_letter].append((c.coordinate, int(m.group(2))))
        print(f"\n=== {sheet} ===")
        bugs = 0
        for col, cells in sorted(by_col.items()):
            rows = [r for _, r in cells]
            modal = Counter(rows).most_common(1)[0][0]
            bad = [(coord, r) for coord, r in cells if r != modal]
            if bad:
                bugs += len(bad)
                for coord, r in bad:
                    print(f"  {coord}: references row {r}, but column {col} should use "
                          f"row {modal}  -> off by {r - modal}")
        if not bugs:
            print("  (no row-reference anomalies)")


if __name__ == "__main__":
    path = sys.argv[1]
    scan(path, ["Merged Set up", "Incoming Set up", "Outgoing Set up"])
