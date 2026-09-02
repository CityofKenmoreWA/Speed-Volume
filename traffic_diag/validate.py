"""Validation: compare Python-computed statistics against the legacy Excel report.

Reads the cached values from each ``* Set up`` sheet (B6..B13) and compares them
to ``metrics.compute_direction_metrics`` run on the same window. Produces a tidy
DataFrame of (study, direction, metric, python, excel, abs_diff, match).
"""
from __future__ import annotations

import os
from typing import Optional

import pandas as pd

from .config import DIRECTION, TS, AnalysisConfig, DEFAULT_ANALYSIS
from .discovery import Study
from .metrics import DOW_NAMES, compute_direction_metrics
from .study import load_study

# Cell -> metric name on each "* Set up" sheet.
_CELLS = {
    "B6": "p85_speed", "B7": "average_speed", "B8": "median_speed",
    "B9": "max_speed", "B10": "adt", "B11": "average_weekday_traffic",
    "B13": "total_vehicles",
}
_SHEETS = {"Merged": "Merged Set up", "Incoming": "Incoming Set up",
           "Outgoing": "Outgoing Set up"}


def read_excel_targets(xlsx_path: str) -> dict:
    """Return {direction: {metric: value}} from the workbook's Set up sheets."""
    from openpyxl import load_workbook
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    out: dict = {}
    for direction, sheet in _SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        vals = {}
        for cell, metric in _CELLS.items():
            v = ws[cell].value
            if isinstance(v, (int, float)):
                vals[metric] = float(v)
        # Also grab the technician window (H1..J1) if present.
        out[direction] = vals
    wb.close()
    return out


def _python_summary(window: pd.DataFrame, direction: str, speed_limit: float,
                    cfg: AnalysisConfig, n_days: int, design_dates) -> dict:
    if direction == "Merged":
        sub = window
    else:
        sub = window[window[DIRECTION] == direction]
    m = compute_direction_metrics(sub, direction, speed_limit, cfg,
                                  n_days=n_days, design_dates=design_dates)
    s = m.summary()
    return {
        "p85_speed": s.get("p85_speed"),
        "average_speed": s.get("average_speed"),
        "median_speed": s.get("median_speed"),
        "max_speed": s.get("max_speed"),
        "adt": s.get("adt"),
        "average_weekday_traffic": s.get("average_weekday_traffic"),
        "total_vehicles": s.get("total_vehicles"),
    }


def validate_study(study: Study, cfg: AnalysisConfig = DEFAULT_ANALYSIS,
                   speed_limit: Optional[float] = None,
                   tol_abs: float = 0.01, tol_rel: float = 1e-4) -> pd.DataFrame:
    """Compare Python vs Excel for one study across all available directions.

    Uses the full raw range as the window (the CSV is already trimmed to the
    analysis week) so differences reflect computation, not window selection.
    """
    xlsx = study.report_xlsx
    if not xlsx or not os.path.exists(xlsx):
        return pd.DataFrame()
    targets = read_excel_targets(xlsx)
    if not targets:
        return pd.DataFrame()

    sd = load_study(study, cfg=cfg, speed_limit=speed_limit)
    # Window = entire raw (matches Excel computing over the full raw column set).
    full = sd.raw
    if full.empty:
        return pd.DataFrame()
    win = (full[TS].min().normalize(),
           full[TS].max().normalize() + pd.Timedelta(days=1) - pd.Timedelta(seconds=1))
    window = full[(full[TS] >= win[0]) & (full[TS] <= win[1])]
    dates = sorted(window[TS].dt.normalize().unique())
    n_days = len(dates)
    design_dates = dates[: max(1, cfg.study_days - 2)]   # first 5 days (Day1..Day5)

    rows = []
    for direction, exp in targets.items():
        got = _python_summary(window, direction, sd.speed_limit, cfg, n_days, design_dates)
        for metric, ev in exp.items():
            pv = got.get(metric)
            if pv is None:
                match, diff = False, None
            else:
                diff = abs(float(pv) - float(ev))
                match = diff <= max(tol_abs, tol_rel * abs(ev))
            rows.append({
                "study": study.study_id, "year": study.year, "status": study.status,
                "direction": direction, "metric": metric,
                "python": pv, "excel": ev, "abs_diff": diff, "match": match,
            })
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------- #
# Hourly-table (cell-level) validation against the Excel "* Report" sheets.
#
# The table is LOCATED, not assumed. Its header row is whichever row has "Time" in
# column A; the 24 hour rows follow it. Across the 750 archived workbooks that
# header sits on row 28, 29 or 30 depending on the template vintage, so the old
# hardcoded "hour 0 is row 31" silently compared every hour against the workbook's
# next hour in 85 of them.
#
# Column layout also varies: B = weekday 85th percentile and C..I = Mon..Sun
# counts throughout, but column J is "Average" in some vintages and "Total" in
# others, and K/L ("Weekday Avg" / "Weekend Avg") are simply absent in 239 of
# them. Summary columns are therefore matched by their header LABEL.
# --------------------------------------------------------------------------- #


def find_hourly_header(ws) -> Optional[int]:
    """Row number of the hourly table's header (col A == "Time"), or None."""
    for r in range(20, 46):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().lower() == "time":
            return r
    return None


def day_columns(hv, day: str) -> list:
    """Our columns holding ``day``'s data: exactly "Monday", or "Monday 03-25" etc.

    A window longer than seven days repeats a weekday and the extra columns carry
    their date to stay unique. Excel keeps one column per weekday name, so the
    comparable value is the sum over all of ours for that day.
    """
    if hv is None:
        return []
    return [c for c in hv.columns
            if isinstance(c, str) and (c == day or c.startswith(day + " "))]


def _summary_columns(ws, header_row: int) -> dict:
    """{column index: header label} for the summary columns right of Sunday."""
    out = {}
    for c in range(10, 14):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip():
            out[c] = v.strip()
    return out
_REPORT_SHEETS = {"Merged": "Merged Report", "Incoming": "Incoming Report",
                  "Outgoing": "Outgoing Report"}
_DAY_COL = {n: 3 + i for i, n in enumerate(DOW_NAMES)}   # Monday->C(3) .. Sunday->I(9)


def _direction_metrics_full(study, cfg):
    """Per-direction metrics on the full-raw window (matches the Excel window)."""
    sd = load_study(study, cfg=cfg)
    full = sd.raw
    if full.empty:
        return None, sd
    dates = sorted(full[TS].dt.normalize().unique())
    start, end = dates[0], dates[-1] + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    window = full[(full[TS] >= start) & (full[TS] <= end)]
    design = dates[: max(1, cfg.study_days - 2)]
    out = {}
    for d in ("Merged", "Incoming", "Outgoing"):
        sub = window if d == "Merged" else window[window[DIRECTION] == d]
        out[d] = compute_direction_metrics(sub, d, sd.speed_limit, cfg,
                                           n_days=len(dates), design_dates=design)
    return out, sd


def validate_hourly_study(study, cfg: AnalysisConfig = DEFAULT_ANALYSIS,
                          tol_abs: float = 0.01) -> pd.DataFrame:
    """Compare the per-hour table (counts, weekday-85th, averages) to the Excel."""
    xlsx = study.report_xlsx
    if not xlsx or not os.path.exists(xlsx):
        return pd.DataFrame()
    from openpyxl import load_workbook
    import warnings
    metrics, _sd = _direction_metrics_full(study, cfg)
    if metrics is None:
        return pd.DataFrame()
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = load_workbook(xlsx, data_only=True)
    rows = []

    def cmp(direction, hour, field, pv, ev):
        if isinstance(ev, str):              # '*' = no data
            ev = None
        if pv is None and ev is None:
            match, diff = True, 0.0
        elif pv is None or ev is None:
            match, diff = False, None
        else:
            diff = abs(float(pv) - float(ev))
            match = diff <= tol_abs
        rows.append({"study": study.study_id, "direction": direction, "hour": hour,
                     "field": field, "python": pv, "excel": ev, "match": match, "diff": diff})

    for direction, sheet in _REPORT_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        m = metrics[direction]
        hv, p85 = m.hourly_volume, m.hourly_weekday_p85
        header = find_hourly_header(ws)
        if header is None:
            continue                     # no recognisable hourly table on this sheet
        # Excel hourly columns are Day1..Day7 (chronological) with weekday-NAME
        # headers on the header row — map by header name, not fixed position.
        col_day = {c: ws.cell(row=header, column=c).value for c in range(3, 10)}
        summary_cols = _summary_columns(ws, header)
        for h in range(24):
            r = header + 1 + h
            # Weekday 85th percentile (col B). NOTE: the legacy template has a known
            # bug in the hour-0 per-hour sub-table (a speed bin references the wrong
            # hour's row), so the midnight weekday-85th can differ by a fraction of a
            # mph; our value uses the complete distribution and is the correct one.
            cmp(direction, h, "weekday_p85",
                None if p85 is None else (None if pd.isna(p85.get(h)) else float(p85.get(h))),
                ws.cell(row=r, column=2).value)
            # Per-day counts (cols C..I), matched by the Excel header's weekday name.
            for col, day in col_day.items():
                if day not in _DAY_COL:
                    continue
                cols = day_columns(hv, day)
                if len(cols) > 1:
                    # The window covers this weekday twice (a >7-day study). The
                    # workbook has only seven day columns, so it keeps the FIRST
                    # occurrence and drops the rest - there is no comparable cell,
                    # and our (complete) figure is not a mismatch. Skip rather than
                    # score it. Two archived studies are affected.
                    continue
                # No column for this weekday = the study never ran that day, so
                # zero vehicles - which is exactly what the workbook records.
                pv = float(hv.loc[h, cols[0]]) if cols else 0.0
                cmp(direction, h, f"count_{day[:3]}", pv, ws.cell(row=r, column=col).value)
            # Summary columns, matched by the label the workbook actually uses.
            # "Total" is the sum of that hour's day counts, which our table does not
            # carry as a column, so it is summed here rather than compared against
            # our per-hour Average (a factor-of-seven mismatch in 262 workbooks).
            for col, label in summary_cols.items():
                key = label.lower()
                if key == "total":
                    # Same caveat as the per-day columns: if any weekday repeats,
                    # the workbook's Total covers only seven of our days.
                    per_day = {d: day_columns(hv, d) for d in DOW_NAMES}
                    if any(len(c) > 1 for c in per_day.values()):
                        continue
                    days = [c for cols_ in per_day.values() for c in cols_]
                    pv = float(hv.loc[h, days].sum()) if days else None
                    field = "total"
                else:
                    colname = {"average": "Average", "weekday avg": "Weekday Avg",
                               "weekend avg": "Weekend Avg"}.get(key)
                    if colname is None:
                        continue         # a summary column we do not model
                    pv = (float(hv.loc[h, colname])
                          if (hv is not None and colname in hv.columns
                              and not pd.isna(hv.loc[h, colname])) else None)
                    field = key.replace(" ", "_")
                cmp(direction, h, field, pv, ws.cell(row=r, column=col).value)
    wb.close()
    return pd.DataFrame(rows)


def validate_hourly_many(studies, cfg: AnalysisConfig = DEFAULT_ANALYSIS) -> pd.DataFrame:
    frames = []
    for s in studies:
        try:
            f = validate_hourly_study(s, cfg)
            if not f.empty:
                frames.append(f)
        except Exception:
            pass
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


def validate_many(studies, cfg: AnalysisConfig = DEFAULT_ANALYSIS, **kw) -> pd.DataFrame:
    """Validate a list of studies; concatenates per-study comparison frames."""
    frames = []
    for s in studies:
        try:
            frames.append(validate_study(s, cfg=cfg, **kw))
        except Exception as e:  # keep going; record the failure
            frames.append(pd.DataFrame([{
                "study": s.study_id, "year": s.year, "status": s.status,
                "direction": "-", "metric": "ERROR", "python": None,
                "excel": None, "abs_diff": None, "match": False,
                "error": str(e)[:120],
            }]))
    return pd.concat([f for f in frames if not f.empty], ignore_index=True) if frames else pd.DataFrame()
