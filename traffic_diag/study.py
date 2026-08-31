"""Load a study's raw data, pick the best complete 7-day window, hold metadata."""
from __future__ import annotations

import os
import warnings
from dataclasses import dataclass, field
from datetime import date
from typing import Optional

import numpy as np
import pandas as pd

from .config import (TS, AnalysisConfig, DEFAULT_ANALYSIS, SourceSpec, SOURCES)
from .discovery import Study, read_notes
from .sources import get_adapter

# Cells (sheet, cell) that hold the posted speed limit in the legacy workbook.
_SPEED_LIMIT_CELLS = [("Merged Set up", "B4"), ("Incoming Set up", "B4"),
                      ("Outgoing Set up", "B4"), ("Merged Report", "B8")]


def read_speed_limit_xlsx(xlsx_path: Optional[str]) -> Optional[float]:
    """Read the posted speed limit from a legacy ``_Report.xlsx`` (None if absent)."""
    if not xlsx_path or not os.path.exists(xlsx_path):
        return None
    try:
        from openpyxl import load_workbook
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            wb = load_workbook(xlsx_path, data_only=True, read_only=True)
        try:
            for sheet, cell in _SPEED_LIMIT_CELLS:
                if sheet in wb.sheetnames:
                    v = wb[sheet][cell].value
                    if isinstance(v, (int, float)) and v > 0:
                        return float(v)
        finally:
            wb.close()
    except Exception:
        return None
    return None


def read_speed_limit(study: Study) -> Optional[float]:
    """Posted speed limit for a study, taken from its existing Excel report."""
    return read_speed_limit_xlsx(study.report_xlsx)


def resolve_speed_limit(study: Study, cfg: AnalysisConfig = DEFAULT_ANALYSIS,
                        explicit: Optional[float] = None,
                        notes: Optional[dict] = None) -> tuple:
    """Resolve the speed limit with documented precedence; returns (value, source).

      1. ``explicit``               — value passed on the CLI / dashboard ("input")
      2. a ``Limit: <n>`` line in ``_Notes.txt``  ("notes")
      3. existing ``_Report.xlsx``  ("excel") — fallback while migrating off Excel
      4. ``cfg.default_speed_limit`` (25 mph)  ("default")
    """
    if explicit is not None:
        return float(explicit), "input"
    if notes is None:
        notes = read_notes(study)
    if notes.get("limit") is not None:
        return float(notes["limit"]), "notes"
    sl = read_speed_limit_xlsx(study.report_xlsx)
    if sl is not None:
        return sl, "excel"
    return float(cfg.default_speed_limit), "default"


@dataclass
class DayCoverage:
    day: date
    count: int
    hours_present: int          # distinct hours 0-23 with >=1 observation
    first_hour: int
    last_hour: int

    @property
    def coverage(self) -> float:
        return self.hours_present / 24.0


@dataclass
class StudyData:
    """Loaded study: full raw frame, selected window frame, metadata + notes."""

    study: Study
    raw: pd.DataFrame
    window: pd.DataFrame
    window_start: pd.Timestamp
    window_end: pd.Timestamp
    speed_limit: float
    speed_limit_source: str = "default"   # input | excel | notes | default
    notes: dict = field(default_factory=dict)
    day_coverage: list = field(default_factory=list)
    selection_note: str = ""

    @property
    def directions(self) -> dict:
        """Canonical->agency direction labels from notes (e.g. Incoming->NB)."""
        return {"Incoming": self.notes.get("incoming"),
                "Outgoing": self.notes.get("outgoing")}

    @property
    def n_days_available(self) -> int:
        return self.raw[TS].dt.normalize().nunique() if not self.raw.empty else 0


def _day_coverage(df: pd.DataFrame) -> list[DayCoverage]:
    if df.empty:
        return []
    g = df.assign(d=df[TS].dt.normalize(), h=df[TS].dt.hour).groupby("d")
    cov = []
    for d, sub in g:
        cov.append(DayCoverage(
            day=d.date(), count=int(len(sub)),
            hours_present=int(sub["h"].nunique()),
            first_hour=int(sub["h"].min()), last_hour=int(sub["h"].max()),
        ))
    return sorted(cov, key=lambda c: c.day)


def select_window(df: pd.DataFrame, cfg: AnalysisConfig = DEFAULT_ANALYSIS):
    """Choose the best complete N-day window.

    Strategy: identify "complete" days (good hourly coverage, dropping partial
    install/removal boundary days), find consecutive runs, then pick the N-day
    block that prefers a Monday start and otherwise maximizes total volume.
    Returns (start_ts, end_ts, note, day_coverage_list).
    """
    cov = _day_coverage(df)
    if not cov:
        raise ValueError("No data to select a window from")
    n_dates = len(cov)
    counts = np.array([c.count for c in cov], dtype=float)
    med = float(np.median(counts)) if len(counts) else 0.0

    # Only the boundary (install / removal) days are treated as partial: an
    # afternoon install leaves the first date covering just the afternoon, and a
    # morning removal leaves the last date covering just the morning. Quiet
    # low-volume interior days are kept (their gaps are a diagnostics concern,
    # not a reason to shrink the window).
    def is_partial_boundary(i: int) -> bool:
        c = cov[i]
        low_vol = med and c.count < 0.5 * med
        if i == 0:
            return c.first_hour > 12 or low_vol
        if i == n_dates - 1:
            return c.last_hour < 12 or low_vol
        return False

    flags = [not is_partial_boundary(i) for i in range(n_dates)]
    if not any(flags):
        flags = [True] * n_dates   # diagnostics will flag the quality problem

    # Longest run of consecutive complete days.
    runs, cur = [], []
    for i, c in enumerate(cov):
        if flags[i] and (not cur or (c.day - cov[i - 1].day).days == 1 and flags[i - 1]):
            cur.append(i)
        else:
            if cur:
                runs.append(cur)
            cur = [i] if flags[i] else []
    if cur:
        runs.append(cur)
    if not runs:
        runs = [list(range(len(cov)))]
    longest = max(runs, key=len)

    n = cfg.study_days
    note = ""
    if len(longest) < n:
        idx0, idx1 = longest[0], longest[-1]
        note = f"Only {len(longest)} complete consecutive days available (< {n})."
    else:
        # All n-day sub-windows within the longest run; prefer Monday start, then volume.
        best = None
        for s in range(0, len(longest) - n + 1):
            block = longest[s:s + n]
            start_day = cov[block[0]].day
            total = sum(cov[i].count for i in block)
            monday_bonus = 1 if start_day.weekday() == 0 else 0
            key = (monday_bonus, total)
            if best is None or key > best[0]:
                best = (key, block)
        block = best[1]
        idx0, idx1 = block[0], block[-1]
        if cov[idx0].day.weekday() != 0:
            note = "No Monday-aligned complete week; chose highest-volume 7-day block."

    start = pd.Timestamp(cov[idx0].day)
    end = pd.Timestamp(cov[idx1].day) + pd.Timedelta(days=1) - pd.Timedelta(seconds=1)
    return start, end, note, cov


def load_study(study: Study, spec: Optional[SourceSpec] = None,
               cfg: AnalysisConfig = DEFAULT_ANALYSIS,
               speed_limit: Optional[float] = None,
               window: Optional[tuple] = None) -> StudyData:
    """Load + window a study. ``window`` overrides auto-selection (start, end)."""
    spec = spec or SOURCES[study.source_name]
    adapter = get_adapter(spec)
    raw = adapter.load(study.path)
    raw = raw.sort_values(TS).reset_index(drop=True)

    notes = read_notes(study)
    # Precedence lives in resolve_speed_limit: input -> Notes "Limit:" -> Excel -> default.
    speed_limit, speed_limit_source = resolve_speed_limit(study, cfg, explicit=speed_limit,
                                                          notes=notes)

    if window is not None:
        start, end = pd.Timestamp(window[0]), pd.Timestamp(window[1])
        sel_note, cov = "Window supplied by caller.", _day_coverage(raw)
    else:
        start, end, sel_note, cov = select_window(raw, cfg)

    win = raw[(raw[TS] >= start) & (raw[TS] <= end)].reset_index(drop=True)
    return StudyData(study=study, raw=raw, window=win, window_start=start, window_end=end,
                     speed_limit=speed_limit, speed_limit_source=speed_limit_source,
                     notes=notes, day_coverage=cov, selection_note=sel_note)
